from aiogram import Bot, types, Dispatcher
from aiogram.enums import ParseMode
from aiogram.filters.command import CommandStart, Command
from aiogram.utils.keyboard import InlineKeyboardBuilder

import logging
import asyncio
import openpyxl

from aiogram import F
from config import TOKEN

# эксель
book = openpyxl.Workbook()
sheet_city = book.active 
sheet_city['A1'] = 0
sheet_city['B1'] = 'id'
sheet_city['C1'] = 'пользователи'
sheet_city['D1'] = 'город'

sheet_piter = book.create_sheet('ПИТЕР')
sheet_piter['A1'] = 0
sheet_piter['B1'] = 'id'
sheet_piter['C1'] = 'пользователи'

sheet_other = book.create_sheet('НЕ ПИТЕР')
sheet_other['A1'] = 0
sheet_other['B1'] = 'id'
sheet_other['C1'] = 'пользователи'
book.save('data.xlsx')

yana_id = 1828681828
my_id = 753125986
#Яна - 1828681828
#я - 753125986

# логирование
logging.basicConfig(level=logging.INFO) 

#переменные для хранения состояния пользователей
waiting_for_answer = set()  # Пользователи, ожидающие выбора "Да" или "Нет"

bot = Bot(token=TOKEN)
dp = Dispatcher()


# Обработка команды /start
@dp.message(CommandStart())
async def process_start_command(message: types.Message):
    #проверяем есть ли юзер пользователя в хуйне waiting_for_answer
    if message.from_user.id not in waiting_for_answer:
        # Добавляем пользователя в список тех, кто должен сделать выбор
        waiting_for_answer.add(message.from_user.id)
        
        #отправляем сообщение
        await message.reply(
            "Привет! Ты из Санкт-Петербурга?", 
            reply_markup=build_keyboard().as_markup(resize_keyboard=True)
        )
    else:
        await message.reply("Ожидается выбор города.", reply_markup=build_keyboard().as_markup(resize_keyboard=True))


@dp.message(Command('help'))
async def process_help_command(message: types.Message):

    #соответсвенно проверяем есть ли юзер в списке ненажавших на кнопку
    if message.from_user.id in waiting_for_answer: 
        await message.reply("Ожидается выбор города.", reply_markup=build_keyboard().as_markup(resize_keyboard=True))
    else:
        await message.reply("Если ты первым решишь загадку - выиграешь приз! Бот работает до 00:00 по МСК.")


@dp.message()
async def handle_messages(message: types.Message):
    
    if message.from_user.id in waiting_for_answer:
        await message.reply("Ожидается выбор города.", reply_markup=build_keyboard().as_markup(resize_keyboard=True))
    elif "тыква" == message.text.lower():
        
        boo = False
        colCity = sheet_city['B']
        for i in colCity:
            if message.from_user.id == i.value:
                print(sheet_city[i.row][3].value)
                if sheet_city[i.row][3].value == 1:
                    boo = True

        if boo:
            Isinlist = False
            colPiter = sheet_piter['B']
            for i in colPiter:
                if message.from_user.id == i.value:
                    Isinlist = True
            
            if Isinlist:
                await message.reply("Вы уже ответили верно!")
            else:
                row = sheet_piter.max_row
                sheet_piter[row + 1][0].value = sheet_piter[row][0].value + 1
                sheet_piter[row + 1][1].value = message.from_user.id
                sheet_piter[row + 1][2].value = message.from_user.username
                book.save('data.xlsx')
                if sheet_piter[row][0].value + 1 < 3:
                    await message.reply("Поздравляю! Ты написал верный ответ самый первый и выиграл билет на вечеринку Gromoween. С тобой свяжутся в ближайшие дни.")
                    await bot.send_message(my_id, message.from_user.username)
                    await bot.send_message(yana_id, message.from_user.username)
                else:
                    await message.reply("Это верный ответ, но ты был не первым. Ты получаешь секретную ссылку на новый выпуск «Громовых мнений» и можешь посмотреть его одним из первых:\nЮтуб: https://youtu.be/-6LcyB3jnm4\nВК: https://vk.com/video-227203258_456239019?list=ln-i716sRpJwEnPaN3Ql0")
        else:
            Isinlist = False
            colOther = sheet_other['B']
            for i in colOther:
                if message.from_user.id == i.value:
                    Isinlist = True
            
            if Isinlist:
                await message.reply("Вы уже ответили верно!")
            else:
                row = sheet_other.max_row
                sheet_other[row + 1][0].value = sheet_other[row][0].value + 1
                sheet_other[row + 1][1].value = message.from_user.id
                sheet_other[row + 1][2].value = message.from_user.username
                book.save('data.xlsx')
                await message.reply("Это верный ответ. Ты получаешь секретную ссылку на новый выпуск «Громовых мнений» и можешь посмотреть его одним из первых: \nЮтуб: https://youtu.be/-6LcyB3jnm4\nВК: https://vk.com/video-227203258_456239019?list=ln-i716sRpJwEnPaN3Ql0")
    else:
        Isinlist = False
        colPiter = sheet_piter['B']
        for i in colPiter:
            if message.from_user.id == i.value:
                Isinlist = True
        colOther = sheet_other['B']
        for i in colOther:
            if message.from_user.id == i.value:
                Isinlist = True

        if Isinlist:
            await message.reply("Вы уже ответили верно!")
        else:
            await message.reply("Н̴̥̈́̎̏̊̚ѐ̴̧̻̫̺̎͆̿͝ͅв̶̭͔̯̇̇е̷͖̟͌̄͐̓͝р̷͋̄̔̆͜͝н̸̭̹̟̾о̶͍̯̋")


@dp.callback_query(F.data == "да")
async def handle_yes(callback: types.CallbackQuery):

    user_id = callback.from_user.id
    # Убираем пользователя из списка ожидания
    if user_id in waiting_for_answer:
        waiting_for_answer.remove(user_id)

    boo = True
    colID = sheet_city['B']
    for i in colID:
        if callback.from_user.id == i.value:
            boo = False


    if boo:
        row = sheet_city.max_row
        sheet_city[row + 1][0].value = sheet_city[row][0].value + 1
        sheet_city[row + 1][1].value = callback.from_user.id
        sheet_city[row + 1][2].value = callback.from_user.username
        sheet_city[row + 1][3].value = 1
        book.save('data.xlsx')
        
        await callback.message.answer("Оранжевый шар на тонкой ножке,\nС зубастой улыбкой, страшной немножко.\nЧто за чудо-овощ, скажи мне, друг?\nОн нужен на празднике ведьм и злюк.")
        #await callback.answer()
    else:
        await callback.message.answer("Город уже выбран.")
        #await callback.answer()

@dp.callback_query(F.data == "нет")
async def handle_no(callback: types.CallbackQuery):

    
    user_id = callback.from_user.id
    # Убираем пользователя из списка ожидания
    if user_id in waiting_for_answer:
        waiting_for_answer.remove(user_id)


    boo = True
    colID = sheet_city['B']
    for i in colID:
        if callback.from_user.id == i.value:
            boo = False

    if boo:
        row = sheet_city.max_row
        sheet_city[row + 1][0].value = sheet_city[row][0].value + 1
        sheet_city[row + 1][1].value = callback.from_user.id
        sheet_city[row + 1][2].value = callback.from_user.username
        sheet_city[row + 1][3].value = 0
        book.save('data.xlsx')

        await callback.message.answer("Оранжевый шар на тонкой ножке, \nС зубастой улыбкой, страшной немножко.\nЧто за чудо-овощ, скажи мне, друг?\nОн нужен на празднике ведьм и злюк.")
        #await callback.answer()
    else:
        await callback.message.answer("Город уже выбран.")
        #await callback.answer()


# билд одного столбца клавиатуры вручную
def build_keyboard():
    builder = InlineKeyboardBuilder()
    builder.row(
        types.InlineKeyboardButton(text="Да", callback_data="да"),
        types.InlineKeyboardButton(text="Нет", callback_data="нет")
    )
    return builder

async def main():
    await dp.start_polling(bot)

if __name__ == '__main__':
    asyncio.run(main())
